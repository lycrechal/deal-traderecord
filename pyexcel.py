#!/usr/bin/env python
# -*- coding: utf-8 -*-

import pandas as pd
from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.writer.excel import ExcelWriter 
from openpyxl.cell import get_column_letter 
import datetime
import time
import os
multiple=int(raw_input("Y(^o^)Yenter the times "))
poundage=int(raw_input("Y(^o^)Yenter the poundage "))




class TradeRecord(object):
    def __init__(self,name,direction,offsetflag,price,number,timedate):
        self.name=name
        self.direction=direction
        self.offsetflag=offsetflag
        self.price=price
        self.number=number
        self.timedate=timedate
class OutPut(object):
    def __init__(self,change,openinterest,avgprice,avgprofit,pingprofit,amount,duokong,poundage,finalfinal):
        self.change=change
        self.openinterest=openinterest
        self.avgprice=avgprice
        self.avgprofit=avgprofit
        self.pingprofit=pingprofit
        self.amount=amount
        self.duokong=duokong
        self.poundage=poundage
        self.finalfinal=finalfinal


class Methods(object):
    def loaddata(self):
        self.OrginDate=[]
        self.daylist=[]
        self.namelist=[]
        self.rows=len(df)
        print "%d" %(self.rows)
        for i in range(self.rows):
            self.OrginDate.append(TradeRecord('',df.direction.values[i],df.offsetflag.values[i],df.price.values[i],df.volume.values[i],df.date.values[i]+" "+str(df.time.values[i])))
            self.daylist.append(df.date.values[i])
            self.namelist.append(df.name.values[i])
        self.daylist=list(set(self.daylist))
        self.namelist=list(set(self.namelist))
        self.OrginDate.sort(key=lambda x:x.timedate.split(':'))
    def loadmin(self):
        self.rows2=len(g_MinData)
        for i in range(self.rows2):
            if g_MinData.datetime.values[i].split(' ')[0] in self.daylist:
                if int(g_MinData.datetime.values[i].split(' ')[1].split(':')[0])<17:
                    if int(g_MinData.datetime.values[i].split(' ')[1].split(':')[0])==9:
                        if int(g_MinData.datetime.values[i].split(' ')[1].split(':')[1])>14:
                            _record = TradeRecord('','','',g_MinData.price.values[i],'',g_MinData.datetime.values[i])
                            self.OrginDate.append(_record)
                    else:
                        _record = TradeRecord('','','',g_MinData.price.values[i],'',g_MinData.datetime.values[i])
                        self.OrginDate.append(_record)


        self.OrginDate.sort(key=lambda x:x.timedate.split(':'))
        self.finalrows=len(self.OrginDate)
    def attendtime(self,k):
        self.number={}
        date1=self.OrginDate[0].timedate
        kkk=0
        ddd=1
        ppp=1
        volume=0
        for i in range(self.finalrows):

            date2=self.OrginDate[i].timedate

            a=time.strptime(date1, "%Y%m%d %H:%M:%S")
            b=time.strptime(date2, "%Y%m%d %H:%M:%S")
            starttime=datetime.datetime(a[0],a[1],a[2],a[3],a[4],a[5])
            endtime=datetime.datetime(b[0],b[1],b[2],b[3],b[4],b[5])
            es=(endtime-starttime).seconds
            if self.OrginDate[i].number==u'':
                self.number[i]=0
            else:
                self.number[i]=self.OrginDate[i].number
            if es==900:
                col = get_column_letter(ppp+1)
                newws6.cell('%s%s'%(col,k+2)).value=int(volume)
                col = get_column_letter(ddd+1)
                newws5.cell('%s%s'%(col,k+2)).value=float(self.fin[i]-self.fin[kkk])
                date1=self.OrginDate[i].timedate
                kkk=i
                ddd=ddd+1
                ppp=ppp+1
                volume=0
            else:
                if es>900:
                    col = get_column_letter(ppp+1) 
                    newws6.cell('%s%s'%(col,k+2)).value=int(volume)
                    col = get_column_letter(ddd+1)
                    newws5.cell('%s%s'%(col,k+2)).value=float(self.fin[i]-self.fin[kkk])

                    date1=self.OrginDate[i].timedate
                    kkk=i
                    ddd=ddd+1
                    ppp=ppp+1
                    volume=0
                else:
                    volume=volume+self.number[i]



        col = get_column_letter(ddd+1)
        newws5.cell('%s%s'%(col,k+2)).value=float(self.fin[self.finalrows-1]-self.fin[kkk])
        col = get_column_letter(ppp+1) 
        newws6.cell('%s%s'%(col,k+2)).value=int(volume)
    def deal(self):
        a={}
        for i in range(self.finalrows):
            if self.OrginDate[i].direction == u'买':
                a[i]=1
            else:
                a[i]=-1
        b={}
        for i in range(self.finalrows):
            if self.OrginDate[i].offsetflag== u'开仓':
                b[i]=1
            else:
               b[i]=-1

        b1={}
        for i in range(self.finalrows):
            if self.OrginDate[i].direction==u'':
                b1[i]=1
            else:
                if self.OrginDate[i].offsetflag == u'开仓':
                    b1[i]=1
                else:
                    b1[i]=-1
        self.n={}
        if self.OrginDate[0].direction==u'':

            self.n[0]=1
        else:
            self.n[0]=a[0]*b[0]
        for i in range(1,self.finalrows):
            if self.OrginDate[i].direction==u'':
                self.n[i]=self.n[i-1]
            else:
                self.n[i]=a[i]*b[i]

        abc={}
        for i in range(self.finalrows):
            if self.OrginDate[i].direction==u'':
                abc[i]=0
            else:
                abc[i]=self.OrginDate[i].number

        self.c={}
        for i in range(self.finalrows):
            if self.OrginDate[i].direction == u'':
                self.c[i]=0
            else:
                self.c[i]=b[i]*self.OrginDate[i].number

        self.pouno={}
        dageo=0
        self.pouno[0]=abs(self.c[0])*poundage
        for i in range(1,self.finalrows):
            dageo=abs(self.c[i])*poundage
            self.pouno[i]=self.pouno[i-1]+dageo


        z=0
        m=0
        j=0
        k=0
        h1={}
        h2={}
        d1={}
        d2={}
        self.d={}
        for i in range(self.finalrows):
            if self.n[i]==1:
                z=z+self.c[i]
                h1[i]=j
                d1[j]=z
                self.d[i]=z
                j=j+1
            else:
                m=self.c[i]+m
                h2[i]=k
                d2[k]=m
                self.d[i]=m
                k=k+1

        v1={}
        v2={}
        self.e={}
        ba1={}
        ba2={}
        r=0
        l=0
        for i in range(self.finalrows):
            if self.n[i]==1:
                if h1[i]==0:
                    ba1[i]=r
                    self.e[i]=self.OrginDate[i].price
                    v1[r]=self.OrginDate[i].price
                    r=r+1
                elif self.d[i]==0:
                    ba1[i]=r
                    self.e[i]=0
                    v1[r]=0
                    r=r+1
                elif d1[h1[i]-1]==0:
                    ba1[i]=r
                    self.e[i]=self.OrginDate[i].price
                    v1[r]=self.OrginDate[i].price
                    r=r+1
                elif b[i]==1:
                    ba1[i]=r
                    self.e[i]=(v1[ba1[i]-1]*d1[h1[i]-1]+self.OrginDate[i].price*self.OrginDate[i].number)/self.d[i]
                    v1[r]=(v1[ba1[i]-1]*d1[h1[i]-1]+self.OrginDate[i].price*self.OrginDate[i].number)/self.d[i]
                    r=r+1
                else:
                    ba1[i]=r
                    self.e[i]=v1[ba1[i]-1]
                    v1[r]=v1[ba1[i]-1]
                    r=r+1
            else:
                if h2[i]==0:
                    ba2[i]=l
                    self.e[i]=self.OrginDate[i].price
                    v2[l]=self.OrginDate[i].price
                    l=l+1
                elif self.d[i]==0:
                    ba2[i]=l
                    self.e[i]=0
                    v2[l]=0
                    l=l+1
                elif d2[h2[i]-1]==0:
                    ba2[i]=l
                    self.e[i]=self.OrginDate[i].price
                    v2[l]=self.OrginDate[i].price
                    l=l+1
                elif b[i]==1:
                    ba2[i]=l
                    self.e[i]=(v2[ba2[i]-1]*d2[h2[i]-1]+self.OrginDate[i].price*self.OrginDate[i].number)/self.d[i]
                    v2[l]=(v2[ba2[i]-1]*d2[h2[i]-1]+self.OrginDate[i].price*self.OrginDate[i].number)/self.d[i]
                    l=l+1
                else:
                    ba2[i]=l
                    self.e[i]=v2[ba2[i]-1]
                    v2[l]=v2[ba2[i]-1]
                    l=l+1

        z=0
        m=0
        j=0
        k=0
        h1={}
        h2={}
        d1={}
        d2={}
        self.d={}
        for i in range(self.finalrows):
            if self.n[i]==1:
                z=z+self.c[i]
                h1[i]=j
                d1[j]=z
                self.d[i]=z
                j=j+1
            else:
                m=self.c[i]+m
                h2[i]=k
                d2[k]=m
                self.d[i]=m
                k=k+1

        self.j={}
        for i in range(self.finalrows):
            if self.n[i]==1:
                if ba1[i]==0:
                    self.j[i]=0
                elif self.d[i]==0 or d1[h1[i]-1]==0:
                    self.j[i]=0
                elif b[i]==1:
                    self.j[i]=(self.OrginDate[i].price-v1[ba1[i]-1])*d1[h1[i]-1]*multiple
                else:
                    self.j[i]=(self.OrginDate[i].price-v1[ba1[i]-1])*self.d[i]*multiple

            else:
                if ba2[i]==0:
                    self.j[i]=0
                elif self.d[i]==0 or d2[h2[i]-1]==0:
                    self.j[i]=0
                elif b[i]==1:
                    self.j[i]=(self.OrginDate[i].price-v2[ba2[i]-1])*d2[h2[i]-1]*(-1)*multiple
                else:
                    self.j[i]=(self.OrginDate[i].price-v2[ba2[i]-1])*self.d[i]*(-1)*multiple

        self.r={}
        o={}
        s={}
        u1={}
        u2={}
        x1=0
        x2=0
        for i in range(self.finalrows):
            if self.n[i]==1:
                if ba1[i]==0:
                    u1[i]=x1
                    self.r[i]=0
                    o[x1]=0
                    x1=x1+1
                elif b1[i]==1:
                    u1[i]=x1
                    self.r[i]=o[u1[i]-1]
                    o[x1]=o[u1[i]-1]
                    x1=x1+1
                else:
                    u1[i]=x1
                    self.r[i]=o[u1[i]-1]+(self.OrginDate[i].price-v1[ba1[i]-1])*self.OrginDate[i].number*multiple
                    o[x1]=o[u1[i]-1]+(self.OrginDate[i].price-v1[ba1[i]-1])*self.OrginDate[i].number*multiple
                    x1=x1+1
            else:
                if ba2[i]==0:
                    u2[i]=x2
                    self.r[i]=0
                    s[x2]=0
                    x2=x2+1
                elif b1[i]==1:
                    u2[i]=x2
                    self.r[i]=s[u2[i]-1]
                    s[x2]=s[u2[i]-1]
                    x2=x2+1
                else:
                    u2[i]=x2
                    self.r[i]=s[u2[i]-1]+(self.OrginDate[i].price-v2[ba2[i]-1])*self.OrginDate[i].number*(-1)*multiple
                    s[x2]=s[u2[i]-1]+(self.OrginDate[i].price-v2[ba2[i]-1])*self.OrginDate[i].number*(-1)*multiple
                    x2=x2+1
        if x1>0:
            self.duo=o[x1-1]
        else:
            self.duo=0
        if x2>0:
            self.kong=s[x2-1]
        else:
            self.kong=0
        sum=0
        self.fin={}

        for i in range(self.finalrows):
            if i==0:
                self.fin[i]=0
            elif self.n[i]==self.n[i-1]:
                self.fin[i]=self.r[i]+self.j[i]+sum

            else:
                sum=self.r[i-1]+self.j[i-1]
                self.fin[i]=self.r[i]+self.j[i]+sum



        self.supfino={}
        for i in range(self.finalrows):
            self.supfino[i]=-self.fin[i]-self.pouno[i]

        net=0
        self.net={}
        for i in range(self.finalrows):
            net=net+self.c[i]*self.n[i]
            self.net[i]=net

        self.num1={}
        for i in range(self.finalrows):
            if i==0:
                if self.net[i]>0:
                    self.num1[i]=self.OrginDate[i].number
                else:
                    self.num1[i]=0
            elif self.net[i]>0:
                if self.net[i]>0 and self.net[i-1]<0:
                    self.num1[i]=self.net[i-1]+self.OrginDate[i].number
                else:
                    self.num1[i]=self.OrginDate[i].number
            else:
                if self.net[i]<0 and self.net[i-1]<=0:
                    self.num1[i]=0
                elif self.net[i]<0 and self.net[i-1]>0:
                    self.num1[i]=self.net[i-1]

                elif self.net[i]==0 and self.net[i-1]<0:
                    self.num1[i]=0
                else:
                    self.num1[i]=self.OrginDate[i].number


        self.num2={}
        for i in range(self.finalrows):
            if i==0:
                if self.net[i]<0:
                    self.num2[i]=self.OrginDate[i].number
                else:
                    self.num2[i]=0
            elif self.net[i]<0:
                if self.net[i]<0 and self.net[i-1]>0:
                    self.num2[i]=-self.net[i-1]+self.OrginDate[i].number
                else:
                    self.num2[i]=self.OrginDate[i].number
            else:
                if self.net[i]>0 and self.net[i-1]>=0:
                    self.num2[i]=0
                elif self.net[i]>0 and self.net[i-1]<0:
                    self.num2[i]=-self.net[i-1]
                elif self.net[i]==0 and self.net[i-1]>0:
                    self.num2[i]=0
                else:
                    self.num2[i]=self.OrginDate[i].number
    def namediff(self,k):
        sum=0
        self.namefin={}
        ppp=2
        diffday=0
        sd=12
        for i in range(self.finalrows):
            if i==0:
                self.namefin[i]=0
            elif self.n[i]==self.n[i-1]:
                self.namefin[i]=self.r[i]+self.j[i]+sum
                if self.OrginDate[i].timedate.split(' ')[0]!=self.OrginDate[i-1].timedate.split(' ')[0]:
                    col = get_column_letter(ppp)
                    colsd=get_column_letter(sd+time.strptime(self.OrginDate[i-1].timedate.split(' ')[0],"%Y%m%d")[6])
                    newwsname.cell('%s%s'%(col,2*k+1)).value=int(self.OrginDate[i-1].timedate.split(' ')[0])
                    newwsname.cell('%s%s'%(col,2*k+2)).value=self.namefin[i]-diffday
                    newws2.cell('%s%d'%(colsd,k+2)).value=self.namefin[i]-diffday
                    diffday=self.namefin[i]
                    ppp=ppp+1
            else:
                sum=self.r[i-1]+self.j[i-1]
                self.namefin[i]=self.r[i]+self.j[i]+sum
                if self.OrginDate[i].timedate.split(' ')[0]!=self.OrginDate[i-1].timedate.split(' ')[0]:
                    col = get_column_letter(ppp)
                    colsd=get_column_letter(sd+time.strptime(self.OrginDate[i-1].timedate.split(' ')[0],"%Y%m%d")[6])
                    newwsname.cell('%s%s'%(col,2*k+1)).value=int(self.OrginDate[i-1].timedate.split(' ')[0])
                    newwsname.cell('%s%s'%(col,2*k+2)).value=self.namefin[i]-diffday
                    newws2.cell('%s%d'%(colsd,k+2)).value=self.namefin[i]-diffday
                    diffday=self.namefin[i]
                    ppp=ppp+1
        col = get_column_letter(ppp)
        colsd=get_column_letter(sd+time.strptime(self.OrginDate[i-1].timedate.split(' ')[0],"%Y%m%d")[6])
        newwsname.cell('%s%s'%(col,2*k+1)).value=int(self.OrginDate[i-1].timedate.split(' ')[0])
        newwsname.cell('%s%s'%(col,2*k+2)).value=self.namefin[i]-diffday
        newws2.cell('%s%d'%(colsd,k+2)).value=self.namefin[i]-diffday

    def spduo(self):

        self.OrginDate1=[]
        for i in range(self.finalrows):
            self.OrginDate1.append(TradeRecord('',self.OrginDate[i].direction,self.OrginDate[i].offsetflag,self.OrginDate[i].price,self.num1[i],self.OrginDate[i].timedate))
        self.OrginDate1.sort(key=lambda x:x.timedate.split(':'))
        
    def func1(self):
        ad={}
        for i in range(self.finalrows):
            if self.OrginDate1[i].direction == u'':
                ad[i]=1
            else:
                if self.OrginDate1[i].direction == u'买':
                    ad[i]=1
                else:
                    ad[i]=-1

        bd={}
        for i in range(self.finalrows):
            if self.OrginDate1[i].offsetflag== u'开仓':
                bd[i]=1
            else:
                bd[i]=-1
        duonumber={}
        for i in range(self.finalrows):
            if self.OrginDate1[i].direction==u'':
                duonumber[i]=0
            else:
                duonumber[i]=self.OrginDate1[i].number
        self.cd={}
        for i in range(self.finalrows):
            if self.OrginDate1[i].direction == u'':
                self.cd[i]=0
            else:
                self.cd[i]=ad[i]*self.OrginDate1[i].number
        self.poundageduo={}
        dage=0
        self.poundageduo[0]=abs(self.cd[0])*poundage
        for i in range(1,self.finalrows):
            dage=abs(self.cd[i])*poundage
            self.poundageduo[i]=self.poundageduo[i-1]+dage
        nd={}

        for i in range(self.finalrows):
            nd[i]=bd[i]*ad[i]

        zad=0
        self.dad={}
        for i in range(self.finalrows):
            zad=zad+self.cd[i]
            self.dad[i]=zad

        self.ed={}
        for i in range(self.finalrows):
            if i==0:
                self.ed[i]=self.OrginDate1[i].price
            elif self.dad[i]==0:
                self.ed[i]=0
            elif self.dad[i-1]==0:
                self.ed[i]=self.OrginDate1[i].price
            elif ad[i]==1:
                self.ed[i]=(self.ed[i-1]*self.dad[i-1]+self.OrginDate1[i].price*duonumber[i])/self.dad[i]
            else:
                self.ed[i]=self.ed[i-1]

        self.jd={}
        for i in range(self.finalrows):
            if i==0:
                self.jd[i]=0
            elif self.dad[i]==0 or self.dad[i-1]==0:
                self.jd[i]=0
            elif ad[i]==1:
                self.jd[i]=(self.OrginDate1[i].price-self.ed[i-1])*self.dad[i-1]*multiple
            else:
                self.jd[i]=(self.OrginDate1[i].price-self.ed[i-1])*self.dad[i-1]*multiple


        self.rd={}
        for i in range(self.finalrows):
            if i==0:
                self.rd[i]=0
            elif ad[i]==1:
                self.rd[i]=self.rd[i-1]
            else:
                self.rd[i]=self.rd[i-1]+(self.OrginDate1[i].price-self.ed[i-1])*self.OrginDate1[i].number*multiple


        self.find={}
        for i in range(self.finalrows):
            if i==0:
                self.find[i]=0
            else:
                self.find[i]=self.rd[i]+self.jd[i]
        self.supfin={}
        for i in range(self.finalrows):
            self.supfin[i]=self.find[i]-self.poundageduo[i]
    def spduokong(self):
        final2=[]
        for i in range(self.finalrows):
            final2.append(OutPut(self.cd[i],'',self.ed[i],self.jd[i],self.rd[i],self.dad[i],'','',self.find[i]))

            newws.cell('G%d'%(i+1)).value=int(final2[i].amount)
            newws.cell('H%d'%(i+1)).value=float(final2[i].finalfinal)
    def netduo(self):
        final2=[]
        for i in range(self.finalrows):
            final2.append(OutPut(self.cd[i],'',self.ed[i],self.jd[i],self.rd[i],-self.find[i],self.dad[i],self.poundageduo[i],-self.find[i]-self.poundageduo[i]))
            newws1.cell('A%d'%(i+1)).value=self.OrginDate1[i].timedate
            newws1.cell('B%d'%(i+1)).value=self.OrginDate1[i].price
            newws1.cell('C%d'%(i+1)).value=int(final2[i].duokong)
            newws1.cell('D%d'%(i+1)).value=float(final2[i].amount)
            newws1.cell('E%d'%(i+1)).value=float(final2[i].poundage)
            newws1.cell('F%d'%(i+1)).value=float(final2[i].finalfinal)

    def spkong(self):

        self.OrginDate2=[]
        for i in range(self.finalrows):
            self.OrginDate2.append(TradeRecord('',self.OrginDate[i].direction,self.OrginDate[i].offsetflag,self.OrginDate[i].price,self.num2[i],self.OrginDate[i].timedate))
        self.OrginDate2.sort(key=lambda x:x.timedate.split(':'))

    def func2(self):
        ak={}
        for i in range(self.finalrows):
            if self.OrginDate2[i].direction == u'':
                ak[i]=-1
            else:
                if self.OrginDate2[i].direction == u'买':
                    ak[i]=1
                else:
                    ak[i]=-1
        bk={}
        for i in range(self.finalrows):
            if self.OrginDate2[i].offsetflag== u'开仓':
                bk[i]=1
            else:
                bk[i]=-1

        self.ck={}
        for i in range(self.finalrows):
            if self.OrginDate2[i].direction == u'':
                self.ck[i]=0
            else:
                self.ck[i]=ak[i]*self.OrginDate2[i].number
        self.poundagekong={}
        dage=0
        self.poundagekong[0]=abs(self.ck[0])*poundage
        for i in range(1,self.rows+self.rows2):
            dage=abs(self.ck[i])*poundage
            self.poundagekong[i]=self.poundagekong[i-1]+dage
        nk={}
        for i in range(self.finalrows):
            nk[i]=bk[i]*ak[i]

        zak=0
        self.dak={}
        for i in range(self.finalrows):
            zak=zak+self.ck[i]
            self.dak[i]=-zak
        kongnumber={}
        for i in range(self.finalrows):
            if self.OrginDate2[i].direction==u'':
                kongnumber[i]=0
            else:
                kongnumber[i]=self.OrginDate2[i].number

        self.ek={}
        for i in range(0,self.finalrows):
            if i==0:
                self.ek[i]=self.OrginDate2[i].price
            elif self.dak[i]==0:
                self.ek[i]=0
            elif self.dak[i-1]==0:
                self.ek[i]=self.OrginDate2[i].price
            elif ak[i]==-1:
                self.ek[i]=(self.ek[i-1]*self.dak[i-1]+self.OrginDate2[i].price*kongnumber[i])/self.dak[i]
            else:
                self.ek[i]=self.ek[i-1]

        self.jk={}
        for i in range(0,self.finalrows):
            if i==0:
                self.jk[i]=0
            elif self.dak[i]==0 or self.dak[i-1]==0:
                self.jk[i]=0
            elif ak[i]==-1:
                self.jk[i]=(self.OrginDate2[i].price-self.ek[i-1])*self.dak[i-1]*multiple*(-1)
            else:
                self.jk[i]=(self.OrginDate2[i].price-self.ek[i-1])*self.dak[i-1]*multiple*(-1)

        self.rk={}
        for i in range(0,self.finalrows):
            if i==0:
                self.rk[i]=0
            elif ak[i]==-1:
                self.rk[i]=self.rk[i-1]
            else:
                self.rk[i]=self.rk[i-1]+(self.OrginDate2[i].price-self.ek[i-1])*self.OrginDate2[i].number*multiple*(-1)


        self.fink={}
        for i in range(0,self.finalrows):
            if i==0:
                self.fink[i]=0
            else:
                self.fink[i]=self.rk[i]+self.jk[i]
        self.supfin={}
        for i in range(self.rows+self.rows2):
            self.supfin[i]=self.fink[i]-self.poundagekong[i]
    def spduokong2(self):
        final3=[]
        for i in range(self.finalrows):
            final3.append(OutPut(self.ck[i],'',self.ek[i],self.jk[i],self.rk[i],self.dak[i],'','',self.fink[i]))

            newws.cell('I%d'%(i+1)).value=int(final3[i].amount)
            newws.cell('J%d'%(i+1)).value=float(final3[i].finalfinal)
    def netkong(self):
        final3=[]
        for i in range(self.finalrows):
            final3.append(OutPut(self.ck[i],'',self.ek[i],self.jk[i],self.rk[i],-self.fink[i],self.dak[i],self.poundagekong[i],-self.fink[i]-self.poundagekong[i]))
            newws2.cell('A%d'%(i+1)).value=self.OrginDate2[i].timedate
            newws2.cell('B%d'%(i+1)).value=self.OrginDate2[i].price
            newws2.cell('C%d'%(i+1)).value=int(final3[i].duokong)
            newws2.cell('D%d'%(i+1)).value=float(final3[i].amount)
            newws2.cell('E%d'%(i+1)).value=float(final3[i].poundage)
            newws2.cell('F%d'%(i+1)).value=float(final3[i].finalfinal)

    def finalresult(self,strfilename):
        final=[]
        ping[strfilename]=self.fin[self.finalrows-1]
        for i in range(self.finalrows):
            final.append(OutPut(self.c[i],self.d[i],self.e[i],self.j[i],self.r[i],self.fin[i],self.net[i],self.pouno[i],self.supfino[i]))

            newws.cell('A%d'%(i+1)).value=self.OrginDate[i].timedate
            newws.cell('B%d'%(i+1)).value=float(self.OrginDate[i].price)
            newws.cell('C%d'%(i+1)).value=int(final[i].duokong)
            newws.cell('D%d'%(i+1)).value=float(final[i].amount)
            newws.cell('E%d'%(i+1)).value=float(final[i].poundage)
            newws.cell('F%d'%(i+1)).value=float(final[i].finalfinal)
    def runtime(self):
        time1=self.OrginDate[0].timedate
        self.alltime=0
        for i in range(self.finalrows):
            if self.net[i]==0:
                time2=self.OrginDate[i].timedate
                a=time.strptime(time1, "%Y%m%d %H:%M:%S")
                b=time.strptime(time2, "%Y%m%d %H:%M:%S")
                starttime=datetime.datetime(a[0],a[1],a[2],a[3],a[4],a[5])
                endtime=datetime.datetime(b[0],b[1],b[2],b[3],b[4],b[5])
                es=(endtime-starttime).seconds
                self.alltime=es+self.alltime
                if i<self.finalrows-1:
                    time1=self.OrginDate[i+1].timedate
                else:
                    time1=time2
    def runduotime(self):
        time1=self.OrginDate[0].timedate
        self.duotime=0
        for i in range(self.finalrows):
            if self.dad[i]==0:
                time2=self.OrginDate[i].timedate
                a=time.strptime(time1, "%Y%m%d %H:%M:%S")
                b=time.strptime(time2, "%Y%m%d %H:%M:%S")
                starttime=datetime.datetime(a[0],a[1],a[2],a[3],a[4],a[5])
                endtime=datetime.datetime(b[0],b[1],b[2],b[3],b[4],b[5])
                es=(endtime-starttime).seconds
                self.duotime=es+self.duotime
                if i<self.finalrows-1:
                    time1=self.OrginDate[i+1].timedate
                else:
                    time1=time2
    def runkongtime(self):
        time1=self.OrginDate[0].timedate
        self.kongtime=0
        for i in range(self.finalrows):
            if self.dak[i]==0:
                time2=self.OrginDate[i].timedate
                a=time.strptime(time1, "%Y%m%d %H:%M:%S")
                b=time.strptime(time2, "%Y%m%d %H:%M:%S")
                starttime=datetime.datetime(a[0],a[1],a[2],a[3],a[4],a[5])
                endtime=datetime.datetime(b[0],b[1],b[2],b[3],b[4],b[5])
                es=(endtime-starttime).seconds
                self.kongtime=es+self.kongtime
                if i<self.finalrows-1:
                    time1=self.OrginDate[i+1].timedate
                else:
                    time1=time2                 
    def everyone(self,k,strfilename,NOD):
        c=0
        duocount=0
        kongcount=0
        newws2.cell('A%d'%(k+2)).value=strfilename
        newws2.cell('B%d'%(k+2)).value=float(ping[strfilename])
        for i in range(self.rows):
            c=c+df.volume.values[i]
        newws2.cell('C%d'%(k+2)).value=int(c)
        newws2.cell('D%d'%(k+2)).value=float(self.duo)
        newws2.cell('E%d'%(k+2)).value=float(self.kong)
        for i in range(self.finalrows):
            if self.n[i]==1 and self.OrginDate[i].number!='':
                duocount=duocount+self.OrginDate[i].number
            elif self.n[i]==-1 and self.OrginDate[i].number!='':
                kongcount=kongcount+self.OrginDate[i].number
        newws2.cell('F%d'%(k+2)).value=int(duocount)
        newws2.cell('G%d'%(k+2)).value=int(kongcount)

        if NOD==0:
            newws2.cell('H%d'%(k+2)).value=int(len(self.daylist))
            newws2.cell('I%d'%(k+2)).value=self.alltime
            newws2.cell('J%d'%(k+2)).value=self.duotime
            newws2.cell('K%d'%(k+2)).value=self.kongtime
            
        else:
            newws2.cell('H%d'%(k+2)).value=int(len(self.namelist))
            newws2.cell('J%d'%(k+2)).value=self.duotime
            newws2.cell('I%d'%(k+2)).value=self.alltime
            newws2.cell('K%d'%(k+2)).value=self.kongtime
            newws2.cell('L%d'%(k+2)).value=time.strptime(strfilename,"%Y%m%d")[6]+1
    def kline(self,k):
        noon=0
        for i in range(self.finalrows):
            if int(self.OrginDate[i].timedate.split(' ')[1].split(':')[0])<12:
                noon=i
                
        if noon==0:
            self.am=[0]
            self.pm=self.supfino.values()
        else:
            self.am=self.supfino.values()[0:noon+1]
            self.pm=self.supfino.values()[noon+1:]
        if len(self.pm)==0:
            self.pm=[0]
            newws3.cell('C%d'%(2*k+2)).value=float(max(self.am))
            newws3.cell('D%d'%(2*k+2)).value=float(min(self.am))
            newws3.cell('E%d'%(2*k+2)).value=float(self.am[-1])
            newws3.cell('C%d'%(2*k+3)).value=0
            newws3.cell('D%d'%(2*k+3)).value=0
            newws3.cell('E%d'%(2*k+3)).value=0
            newws3.cell('B%d'%(2*k+2)).value=0
            newws3.cell('B%d'%(2*k+3)).value=0
            #newws3.cell('J%d'%(k+2)).value=max(self.supfino.items(), key=lambda x: x[1])[1]
            #newws3.cell('K%d'%(k+2)).value=min(self.supfino.items(), key=lambda x: x[1])[1]
            #newws3.cell('L%d'%(k+2)).value=self.supfino[max(self.supfino)]
            #newws3.cell('I%d'%(k+2)).value=0
            newws3.cell('H%d'%(2*k+2)).value=float((max(self.am))/(len(self.namelist)))
            newws3.cell('I%d'%(2*k+2)).value=float((min(self.am))/(len(self.namelist)))
            newws3.cell('J%d'%(2*k+2)).value=float(self.am[-1]/(len(self.namelist)))
            newws3.cell('G%d'%(2*k+2)).value=0
            newws3.cell('H%d'%(2*k+3)).value=0
            newws3.cell('I%d'%(2*k+3)).value=0
            newws3.cell('J%d'%(2*k+3)).value=0
            newws3.cell('G%d'%(2*k+3)).value=0
        else:
        #newws3.cell('F%d'%(2*k+2)).value=max(self.am)
        #newws3.cell('G%d'%(2*k+2)).value=min(self.am)
        #newws3.cell('H%d'%(2*k+2)).value=self.am[-1]
        #newws3.cell('E%d'%(2*k+3)).value=self.am[-1]
        #newws3.cell('F%d'%(2*k+3)).value=max(self.pm)
        #newws3.cell('G%d'%(2*k+3)).value=min(self.pm)
        #newws3.cell('H%d'%(2*k+3)).value=self.pm[-1]
        #newws3.cell('E%d'%(2*k+4)).value=self.pm[-1]
        #newws3.cell('B%d'%(k+2)).value=max(self.supfino.items(), key=lambda x: x[1])[1]
        #newws3.cell('C%d'%(k+2)).value=min(self.supfino.items(), key=lambda x: x[1])[1]
        #newws3.cell('D%d'%(k+2)).value=self.supfino[max(self.supfino)]
        #newws3.cell('A%d'%(k+3)).value=self.supfino[max(self.supfino)]
            #newws3.cell('J%d'%(k+2)).value=max(self.supfino.items(), key=lambda x: x[1])[1]
            #newws3.cell('K%d'%(k+2)).value=min(self.supfino.items(), key=lambda x: x[1])[1]
            #newws3.cell('L%d'%(k+2)).value=self.supfino[max(self.supfino)]
            #newws3.cell('I%d'%(k+2)).value=0
            newws3.cell('C%d'%(2*k+2)).value=float(max(self.am))
            newws3.cell('D%d'%(2*k+2)).value=float(min(self.am))
            newws3.cell('E%d'%(2*k+2)).value=float(self.am[-1])
            newws3.cell('B%d'%(2*k+2)).value=0
            newws3.cell('C%d'%(2*k+3)).value=float(max(self.pm)-self.am[-1])
            newws3.cell('D%d'%(2*k+3)).value=float(min(self.pm)-self.am[-1])
            newws3.cell('E%d'%(2*k+3)).value=float(self.pm[-1]-self.am[-1])
            newws3.cell('B%d'%(2*k+3)).value=0
            newws3.cell('H%d'%(2*k+2)).value=float((max(self.am))/(len(self.namelist)))
            newws3.cell('I%d'%(2*k+2)).value=float((min(self.am))/(len(self.namelist)))
            newws3.cell('J%d'%(2*k+2)).value=float(self.am[-1]/(len(self.namelist)))
            newws3.cell('G%d'%(2*k+2)).value=0
            newws3.cell('H%d'%(2*k+3)).value=float((max(self.pm)-self.am[-1])/(len(self.namelist)))
            newws3.cell('I%d'%(2*k+3)).value=float((min(self.pm)-self.am[-1])/(len(self.namelist)))
            newws3.cell('J%d'%(2*k+3)).value=float((self.pm[-1]-self.am[-1])/(len(self.namelist)))
            newws3.cell('G%d'%(2*k+3)).value=0
        newws3.cell('A%d'%(2*k+2)).value=strfilename+'A'
        newws3.cell('A%d'%(2*k+3)).value=strfilename+'B'
if __name__=='__main__':

    g_TradeData=pd.read_excel('C://Users//Administrator//result//all.xlsx', 'Sheet1', index_col=None, na_values=['NA'],converters={'date':str,'price':float})
    g_MinData=pd.read_excel('C://Users//Administrator//result//min.xls', 'Sheet1', index_col=None, na_values=['NA'])

    mydist=sorted(list(set(g_TradeData['date'])))
    print mydist
    calculatedate=Methods()
    newwb=Workbook()
    ew = ExcelWriter(workbook = newwb)
    dest_filename = r'C://Users//Administrator//result//out.xlsx'
    newws2=newwb.create_sheet(u'结算单（日期）')
    newws3=newwb.create_sheet(u'开低高收')
    newws2.cell('A1').value=u'日期'
    newws2.cell('B1').value=u'平仓盈亏'
    newws2.cell('C1').value=u'成交量'
    newws2.cell('D1').value=u'平仓盈亏多头'
    newws2.cell('E1').value=u'平仓盈亏空头'
    newws2.cell('F1').value=u'成交量多头'
    newws2.cell('G1').value=u'成交量空头'
    newws2.cell('H1').value=u'交易人数'
    newws2.cell('I1').value=u'持仓时间'
    newws2.cell('I1').value=u'持仓时间'
    newws2.cell('J1').value=u'多头持仓时间'
    newws2.cell('K1').value=u'空头持仓时间'
    newws2.cell('L1').value=u'周几'
    ping={}
    newws5=newwb.create_sheet(u'时间段结算（盈亏）')
    newws6=newwb.create_sheet(u'时间段结算（成交量）')
    newws5.cell('A1').value=u'日期'
    newws6.cell('A1').value=u'日期'
    newws5.cell('B1').value=u'09:15:00-09:30:00'
    newws6.cell('B1').value=u'09:15:00-09:30:00'
    newws5.cell('C1').value=u'09:30:00-09:45:00'
    newws6.cell('C1').value=u'09:30:00-09:45:00'
    newws5.cell('D1').value=u'09:45:00-10:00:00'
    newws6.cell('D1').value=u'09:45:00-10:00:00'
    newws5.cell('E1').value=u'10:00:00-10:15:00'
    newws6.cell('E1').value=u'10:00:00-10:15:00'
    newws5.cell('F1').value=u'10:15:00-10:30:00'
    newws6.cell('F1').value=u'10:15:00-10:30:00'
    newws5.cell('G1').value=u'10:30:00-10:45:00'
    newws6.cell('G1').value=u'10:30:00-10:45:00'
    newws5.cell('H1').value=u'10:45:00-11:00:00'
    newws6.cell('H1').value=u'10:45:00-11:00:00'
    newws5.cell('I1').value=u'11:00:00-11:15:00'
    newws6.cell('I1').value=u'11:00:00-11:15:00'
    newws5.cell('J1').value=u'11:15:00-11:30:00'
    newws6.cell('J1').value=u'11:15:00-11:30:00'
    newws5.cell('K1').value=u'11:30:00-11:45:00'
    newws6.cell('K1').value=u'11:30:00-11:45:00'
    newws5.cell('L1').value=u'11:45:00-12:00:00'
    newws6.cell('L1').value=u'11:45:00-12:00:00'
    newws5.cell('M1').value=u'13:00:00-13:15:00'
    newws6.cell('M1').value=u'13:00:00-13:15:00'
    newws5.cell('N1').value=u'13:15:00-13:30:00'
    newws6.cell('N1').value=u'13:15:00-13:30:00'
    newws5.cell('O1').value=u'13:30:00-13:45:00'
    newws6.cell('O1').value=u'13:30:00-13:45:00'
    newws5.cell('P1').value=u'13:45:00-14:00:00'
    newws6.cell('P1').value=u'13:45:00-14:00:00'
    newws5.cell('Q1').value=u'14:00:00-14:15:00'
    newws6.cell('Q1').value=u'14:00:00-14:15:00'
    newws5.cell('R1').value=u'14:15:00-14:30:00'
    newws6.cell('R1').value=u'14:15:00-14:30:00'
    newws5.cell('S1').value=u'14:30:00-14:45:00'
    newws6.cell('S1').value=u'14:30:00-14:45:00'
    newws5.cell('T1').value=u'14:45:00-15:00:00'
    newws6.cell('T1').value=u'14:45:00-15:00:00'
    newws5.cell('U1').value=u'15:00:00-15:15:00'
    newws6.cell('U1').value=u'15:00:00-15:15:00'
    newws5.cell('V1').value=u'15:15:00-15:30:00'
    newws6.cell('V1').value=u'15:15:00-15:30:00'
    newws5.cell('W1').value=u'15:30:00-15:45:00'
    newws6.cell('W1').value=u'15:30:00-15:45:00'
    newws5.cell('X1').value=u'15:45:00-16:00:00'
    newws6.cell('X1').value=u'15:45:00-16:00:00'
    newws5.cell('Y1').value=u'16:00:00-16:15:00'
    newws6.cell('Y1').value=u'16:00:00-16:15:00'
    for k in range(len(mydist)):
        df=g_TradeData[(g_TradeData['date']==mydist[k])].loc[:,['name','direction','offsetflag','price','volume','time','date']]
        strfilename=str(mydist[k])
        calculatedate.loaddata()
        calculatedate.loadmin()
        newws=newwb.create_sheet(mydist[k])
        newws5.cell('A%d'%(k+2)).value=mydist[k]
        newws6.cell('A%d'%(k+2)).value=mydist[k]
        calculatedate.deal()
        calculatedate.finalresult(strfilename)
        calculatedate.spduo()
        calculatedate.func1()
        calculatedate.spduokong()
        calculatedate.spkong()
        calculatedate.func2()
        calculatedate.spduokong2()
        calculatedate.kline(k)
        calculatedate.runtime()
        calculatedate.runduotime()
        calculatedate.runkongtime()
        calculatedate.everyone(k,strfilename,2)

        calculatedate.attendtime(k)
    newws3.cell('B%d'%(2*len(mydist)+2)).value=''
    newws3.cell('G%d'%(2*len(mydist)+2)).value=''
    ew.save(filename = dest_filename)
    calculateall=Methods()
    newwb=load_workbook(r'C://Users//Administrator//result//out.xlsx')
    ew = ExcelWriter(workbook = newwb)
    dest_filename = r'C://Users//Administrator//result//out.xlsx'
    df=g_TradeData.loc[:,['name','direction','offsetflag','price','volume','time','date']]
    ping={}

    calculateall.loaddata()
    calculateall.loadmin()
    newws=newwb.create_sheet(u'总表')
    calculateall.deal()
    calculateall.finalresult('k')
    calculateall.spduo()
    calculateall.func1()
    calculateall.spduokong()
    calculateall.spkong()
    calculateall.func2()
    calculateall.spduokong2()
    newws1=newwb.create_sheet(u'净多')
    newws2=newwb.create_sheet(u'净空')
    calculateall.spduo()
    calculateall.func1()
    calculateall.netduo()
    calculateall.spkong()
    calculateall.func2()
    calculateall.netkong()
    ew.save(filename = dest_filename)
    mylist=list(set(g_TradeData['name']))
    print mylist
    calculatename=Methods()
    newwb=load_workbook(r'C://Users//Administrator//result//out.xlsx')
    ew = ExcelWriter(workbook = newwb)
    dest_filename = r'C://Users//Administrator//result//out.xlsx'
    df=g_TradeData.loc[:,['name','direction','offsetflag','price','volume','time','date']]
    ping={}
    newws2=newwb.create_sheet(u'结算单（姓名）')
    newws2.cell('A1').value=u'日期'
    newws2.cell('B1').value=u'平仓盈亏'
    newws2.cell('C1').value=u'成交量'
    newws2.cell('D1').value=u'平仓盈亏多头'
    newws2.cell('E1').value=u'平仓盈亏空头'
    newws2.cell('F1').value=u'成交量多头'
    newws2.cell('G1').value=u'成交量空头'
    newws2.cell('H1').value=u'交易天数'
    newws2.cell('I1').value=u'持仓时间'
    newws2.cell('J1').value=u'多头持仓时间'
    newws2.cell('K1').value=u'空头持仓时间'
    newws2.cell('L1').value=u'周一'
    newws2.cell('M1').value=u'周二'
    newws2.cell('N1').value=u'周三'
    newws2.cell('O1').value=u'周四'
    newws2.cell('P1').value=u'周五'
    newwsname=newwb.create_sheet(u'namedaydiff')
    for k in range(len(mylist)):
        df=g_TradeData[(g_TradeData['name']==mylist[k])].loc[:,['name','direction','offsetflag','price','volume','time','date']]
        strfilename=mylist[k]
        calculatename.loaddata()
        calculatename.loadmin()
        newws=newwb.create_sheet(mylist[k])
        newwsname.cell('A%d'%(2*k+2)).value=strfilename
        calculatename.deal()
        calculatename.namediff(k)
        calculatename.finalresult(strfilename)
        calculatename.spduo()
        calculatename.func1()
        calculatename.spduokong()
        calculatename.spkong()
        calculatename.func2()
        calculatename.spduokong2()
        calculatename.runtime()
        calculatename.runduotime()
        calculatename.runkongtime()
        calculatename.everyone(k,strfilename,0)



    ew.save(filename = dest_filename)
